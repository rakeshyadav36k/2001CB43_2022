

# from datetime import datetime
# start_time = datetime.now()

# #Help https://youtu.be/N6PBd4XdnEw
# def octant_range_names(mod=5000):

    
#     octant_id = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}

# ###Code

# from platform import python_version
# ver = python_version()

# if ver == "3.8.10":
#     print("Correct Version Installed")
# else:
#     print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")


# mod=5000 
# octant_range_names(mod)



# #This shall be the last lines of the code.
# end_time = datetime.now()
# print('Duration of Program Execution: {}'.format(end_time - start_time))


from openpyxl import load_workbook
import numpy as np
import pandas as pd

wb = load_workbook('octant_input.xlsx')     # I have opened the Excel file
sheet = wb.active

df = pd.read_excel('octant_input.xlsx')     # df is a data frame in which we put the data of excel file using pandas

sheet.cell(row=1, column=5).value = "U Avg"       # Written the header 
sheet.cell(row=1, column=6).value = "V Avg" 
sheet.cell(row=1, column=7).value = "W Avg"  
sheet.cell(row=1, column=8).value = "U'=U - U avg" 
sheet.cell(row=1, column=9).value = "V'=V - V avg" 
sheet.cell(row=1, column=10).value = "W'=W - W avg"

Uavg = df['U'].mean()     # calculated the mean using pandas
Vavg = df['V'].mean()
Wavg = df['W'].mean()

sheet.cell(row=2, column=5).value = Uavg   # added the average values in the sheet
sheet.cell(row=2, column=6).value = Vavg 
sheet.cell(row=2, column=7).value = Wavg 

l1 = df['U']            # creating the list l1,l2,l3 which consist the element of U,V,W respectively
l2 = df['V']
l3 = df['W']


# creating three lists l4,l5 & l6 which contains the values of U', V' & W'

#********************   
l4=[]
for i in l1:
    a = i - Uavg
    l4.append(a)

for i in range(2,len(l1)+2):
    sheet.cell(row=i, column=8).value = l4[i-2]

l5=[]
for i in l2:
    a = i - Vavg
    l5.append(a)

for i in range(2,len(l2)+2):
    sheet.cell(row=i, column=9).value = l5[i-2]

l6=[]
for i in l3:
    a = i - Wavg
    l6.append(a)

for i in range(2,len(l3)+2):
    sheet.cell(row=i, column=10).value = l6[i-2]

#************************************

# here we have created a header "Octant" and print the values of octants in excel file

sheet.cell(row=1, column=11).value = "Octant"



for i in range(0,len(l1)):
    if(l4[i]>0 and l5[i]>0):
        if(l6[i]>0):
            sheet.cell(row=i+2, column=11).value = "+1"
        else:                                               # this tells whether the octant is +1 or -1
            sheet.cell(row=i+2, column=11).value = "-1"
    elif(l4[i]<0 and l5[i]>0):
        if(l6[i]>0):
            sheet.cell(row=i+2, column=11).value = "+2"
        else:                                               # this tells whether the octant is +2 or -2
            sheet.cell(row=i+2, column=11).value = "-2"
    elif(l4[i]<0 and l5[i]<0):
        if(l6[i]>0):
            sheet.cell(row=i+2, column=11).value = "+3"
        else:                                                # this tells whether the octant is +3 or -3
            sheet.cell(row=i+2, column=11).value = "-3"
    elif(l4[i]>0 and l5[i]<0):
        if(l6[i]>0):
            sheet.cell(row=i+2, column=11).value = "+4"
        else:                                                 # this tells whether the octant is +4 or -4
            sheet.cell(row=i+2, column=11).value = "-4"
    
        
# l7 = df['Octant']  # this list l7 contains the all octants values
l7=[]
for i in range(len(l1)):
    x=sheet.cell(row=i+2,column=11).value
    l7.append(int(x))


sheet['M2']="Overall Count"  # this is basicallly printed the header
sheet['N1'] = "+1"
sheet['O1'] = "-1"
sheet['P1'] = "+2"
sheet['Q1'] = "-2"
sheet['R1'] = "+3"
sheet['S1'] = "-3"
sheet['T1'] = "+4"
sheet['U1'] = "-4"

ctpos1 = ctneg1 = ctpos2 = ctneg2 = ctpos3 = ctneg3 = ctpos4 = ctneg4 = 0  # these variables are total no each octant present

for i in range(0,len(l1)):
    if(l4[i]>0 and l5[i]>0):
        if(l6[i]>0):
            ctpos1 += 1               # total count of octant no +1 & -1
        else:
            ctneg1 += 1
    elif(l4[i]<0 and l5[i]>0):
        if(l6[i]>0):
            ctpos2 += 1                 # total count of octant no +2 & -2
        else:
            ctneg2 += 1
    elif(l4[i]<0 and l5[i]<0):
        if(l6[i]>0):
            ctpos3 += 1                 # total count of octant no +3 & -3
        else:
            ctneg3 += 1
    elif(l4[i]>0 and l5[i]<0):
        if(l6[i]>0):
            ctpos4 += 1                 # total count of octant no +4 & -4
        else:
            ctneg4 += 1


#  we have inserted the values of total no of each count

sheet.cell(row=2, column=14).value = ctpos1
sheet.cell(row=2, column=15).value = ctneg1
sheet.cell(row=2, column=16).value = ctpos2
sheet.cell(row=2, column=17).value = ctneg2
sheet.cell(row=2, column=18).value = ctpos3
sheet.cell(row=2, column=19).value = ctneg3
sheet.cell(row=2, column=20).value = ctpos4
sheet.cell(row=2, column=21).value = ctneg4


mod=5000    # this is a user defined mod value
if(len(l7)%mod!=0):           
    p = len(l7)//mod + 1           #variable p is no of partitions
else:
    p = len(l7)//mod 

sheet['L3'] = "User Input"
sheet['M3'] = "Mod" + " " +str(mod)

A = []              # here we have taken a list A which contains another list B(list B contains the octants values of partition)
x=0
new_mod = mod
for i in range(p):
    B = []
    for j in range(x,x + new_mod):
        B.append(l7[j])
    x+=mod
    if((len(l7)-x)<mod):
            new_mod = len(l7) - x
    A.append(B)


new_octant = [1,-1,2,-2,3,-3,4,-4]   #this is a list of octant values
for i in range(p):
    if(mod*(i+1)<len(l7)):
        sheet.cell(row=i+4, column=13).value = str(mod*i)+"-"+str(mod*(i+1)-1)
    else:
        sheet.cell(row=i+4, column=13).value = str(mod*i)+"-"+str(len(l7)-1)
    for j in range(8):
        sheet.cell(row=i+4, column=14+j).value = A[i].count(new_octant[j])      #we have counted the octant values in a range mod value

for i in range(8):
    sheet.cell(row=1,column=22+i).value = "rank of "+str(new_octant[i])


dictionary_ct={1:ctpos1,-1:ctneg1,2:ctpos2,-2:ctneg2,3:ctpos3,-3:ctneg3,4:ctpos4,-4:ctneg4}
dictionary_ct=dict(sorted(dictionary_ct.items(), key=lambda item:item[1]))
dictionary_ct=list(dictionary_ct.items())

for i in range(8):
    if(dictionary_ct[i][0]==1):
        sheet.cell(row=2,column=22).value=8-i
    elif(dictionary_ct[i][0]==-1):
        sheet.cell(row=2,column=23).value=8-i
    elif(dictionary_ct[i][0]==2):
        sheet.cell(row=2,column=24).value=8-i
    elif(dictionary_ct[i][0]==-2):
        sheet.cell(row=2,column=25).value=8-i
    elif(dictionary_ct[i][0]==3):
        sheet.cell(row=2,column=26).value=8-i
    elif(dictionary_ct[i][0]==-3):
        sheet.cell(row=2,column=27).value=8-i
    elif(dictionary_ct[i][0]==4):
        sheet.cell(row=2,column=28).value=8-i
    elif(dictionary_ct[i][0]==-4):
        sheet.cell(row=2,column=29).value=8-i

sheet['AD1'] = "Rank1 Octant ID"
sheet['AE1'] = "Rank1 Octant Name"

octant_id = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}
sheet['AD2']=dictionary_ct[7][0]
sheet['AE2']=octant_id[str(dictionary_ct[7][0])]

rank1=[]    # this list stores the number of rank 1 octants

for i in range(p):   # made a dictionary to store the count value of each octants
    mod_dictionary_ct={1:A[i].count(1), -1:A[i].count(-1),2:A[i].count(2),-2:A[i].count(-2),3:A[i].count(3),-3:A[i].count(-3),4:A[i].count(4),-4:A[i].count(-4)}
    mod_dictionary_ct=dict(sorted(mod_dictionary_ct.items(),key=lambda item:item[1]))     # sorted the values in increasing order
    mod_dictionary_ct=list(mod_dictionary_ct.items())    # and make a list of that dictionary
    for j in range(8):
        if(mod_dictionary_ct[j][0]==1):
            sheet.cell(row=4+i,column=22).value=8-j
        elif(mod_dictionary_ct[j][0]==-1):
            sheet.cell(row=4+i,column=23).value=8-j
        elif(mod_dictionary_ct[j][0]==2):
            sheet.cell(row=4+i,column=24).value=8-j
        elif(mod_dictionary_ct[j][0]==-2):
            sheet.cell(row=4+i,column=25).value=8-j
        elif(mod_dictionary_ct[j][0]==3):
            sheet.cell(row=4+i,column=26).value=8-j
        elif(mod_dictionary_ct[j][0]==-3):
            sheet.cell(row=4+i,column=27).value=8-j
        elif(mod_dictionary_ct[j][0]==4):
            sheet.cell(row=4+i,column=28).value=8-j
        elif(mod_dictionary_ct[j][0]==-4):
            sheet.cell(row=4+i,column=29).value=8-j

    sheet.cell(row=4+i,column=30).value=mod_dictionary_ct[7][0]
    rank1.append(mod_dictionary_ct[7][0])                   
    sheet.cell(row=4+i,column=31).value=octant_id[str(mod_dictionary_ct[7][0])]

# code to print the octant which 1 rank appear most

sheet.cell(row=14,column=14).value="Octant ID"
sheet.cell(row=14,column=15).value="Octant Name"
sheet.cell(row=14,column=16).value="Count of Rank 1 Mod 'Values"

for i in range(8):                           
    sheet.cell(row=14+i,column=14).value=new_octant[i]
    sheet.cell(row=14+i,column=15).value=octant_id[str(new_octant[i])]
    sheet.cell(row=14+i,column=16).value=rank1.count(new_octant[i])


wb.save("octant_output_ranking_excel_2001CB43.xlsx")


    



