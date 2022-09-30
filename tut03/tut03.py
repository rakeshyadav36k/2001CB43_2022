# #Help https://youtu.be/H37f_x4wAC0
# def octant_longest_subsequence_count():
# ###Code

# from platform import python_version
# ver = python_version()

# if ver == "3.8.10":
#     print("Correct Version Installed")
# else:
#     print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")


# octant_longest_subsequence_count()


# I have imported some libraries

from openpyxl import load_workbook
import numpy as np
import pandas as pd

wb = load_workbook('input_octant_longest_subsequence.xlsx')     # I have opened the Excel file
sheet = wb.active

df = pd.read_excel('input_octant_longest_subsequence.xlsx')     # df is a data frame in which we put the data of excel file using pandas

sheet.cell(row=1, column=5).value = "U Avg"       # Written the header 
sheet.cell(row=1, column=6).value = "V Avg" 
sheet.cell(row=1, column=7).value = "W Avg"  
sheet.cell(row=1, column=8).value = "U'=U - U avg" 
sheet.cell(row=1, column=9).value = "V'=V - V avg" 
sheet.cell(row=1, column=10).value = "W'=W - W avg"

Uavg = df['U'].mean()     # calculated the mean using pandas
Vavg = df['V'].mean()
Wavg = df['W'].mean()

sheet.cell(row=2, column=5).value = Uavg   # added the average values in the sheet
sheet.cell(row=2, column=6).value = Vavg 
sheet.cell(row=2, column=7).value = Wavg 

l1 = df['U']            # creating the list l1,l2,l3 which consist the element of U,V,W respectively
l2 = df['V']
l3 = df['W']


# creating three lists l4,l5 & l6 which contains the values of U', V' & W'

#********************   
l4=[]
for i in l1:
    a = i - Uavg
    l4.append(a)

for i in range(2,len(l1)+2):
    sheet.cell(row=i, column=8).value = l4[i-2]

l5=[]
for i in l2:
    a = i - Vavg
    l5.append(a)

for i in range(2,len(l2)+2):
    sheet.cell(row=i, column=9).value = l5[i-2]

l6=[]
for i in l3:
    a = i - Wavg
    l6.append(a)

for i in range(2,len(l3)+2):
    sheet.cell(row=i, column=10).value = l6[i-2]

#************************************

# here we have created a header "Octant" and print the values of octants in excel file

sheet.cell(row=1, column=11).value = "Octant"



for i in range(0,len(l1)):
    if(l4[i]>0 and l5[i]>0):
        if(l6[i]>0):
            sheet.cell(row=i+2, column=11).value = "+1"
        else:                                               # this tells whether the octant is +1 or -1
            sheet.cell(row=i+2, column=11).value = "-1"
    elif(l4[i]<0 and l5[i]>0):
        if(l6[i]>0):
            sheet.cell(row=i+2, column=11).value = "+2"
        else:                                               # this tells whether the octant is +2 or -2
            sheet.cell(row=i+2, column=11).value = "-2"
    elif(l4[i]<0 and l5[i]<0):
        if(l6[i]>0):
            sheet.cell(row=i+2, column=11).value = "+3"
        else:                                                # this tells whether the octant is +3 or -3
            sheet.cell(row=i+2, column=11).value = "-3"
    elif(l4[i]>0 and l5[i]<0):
        if(l6[i]>0):
            sheet.cell(row=i+2, column=11).value = "+4"
        else:                                                 # this tells whether the octant is +4 or -4
            sheet.cell(row=i+2, column=11).value = "-4"
    

sheet.cell(row=1,column=13).value = "Octants"
sheet.cell(row=1,column=14).value = "Longest Sequence Length"
sheet.cell(row=1,column=15).value = "Count"

list_octant = ["+1","-1","+2","-2","+3","-3","+4","-4"]
for i in range(8):
    sheet.cell(row=2+i,column=13).value = list_octant[i]

octs=[]           # we put all the octants values in list octs
for i in range(len(l1)):
    x=sheet.cell(row=i+2,column=11).value
    octs.append(int(x))


list1=[]      #list containing length of subsequents
count1=1
for i in range(len(octs)-1):
    if(octs[i]==1 and octs[i+1]==1):
        count1+=1
    elif(octs[i]==1 and octs[i+1]!=1):
        list1.append(count1)
        count1=1

list_1=[]      #list containing length of subsequents
count_1=1
for i in range(len(octs)-1):
    if(octs[i]==-1 and octs[i+1]==-1):
        count_1+=1
    elif(octs[i]==-1 and octs[i+1]!=-1):
        list_1.append(count_1)
        count_1=1

list2=[]      #list containing length of subsequents
count2=1
for i in range(len(octs)-1):
    if(octs[i]==2 and octs[i+1]==2):
        count2+=1
    elif(octs[i]==2 and octs[i+1]!=2):
        list2.append(count2)
        count2=1

list_2=[]      #list containing length of subsequents
count_2=1
for i in range(len(octs)-1):
    if(octs[i]==-2 and octs[i+1]==-2):
        count_2+=1
    elif(octs[i]==-2 and octs[i+1]!=-2):
        list_2.append(count_2)
        count_2=1

list3=[]      #list containing length of subsequents
count3=1
for i in range(len(octs)-1):
    if(octs[i]==3 and octs[i+1]==3):
        count3+=1
    elif(octs[i]==3 and octs[i+1]!=3):
        list3.append(count3)
        count3=1

list_3=[]      #list containing length of subsequents
count_3=1
for i in range(len(octs)-1):
    if(octs[i]==-3 and octs[i+1]==-3):
        count_3+=1
    elif(octs[i]==-3 and octs[i+1]!=-3):
        list_3.append(count_3)
        count_3=1

list4=[]      #list containing length of subsequents
count4=1
for i in range(len(octs)-1):
    if(octs[i]==4 and octs[i+1]==4):
        count4+=1
    elif(octs[i]==4 and octs[i+1]!=4):
        list4.append(count4)
        count4=1

list_4=[]      #list containing length of subsequents
count_4=1
for i in range(len(octs)-1):
    if(octs[i]==-4 and octs[i+1]==-4):
        count_4+=1
    elif(octs[i]==-4 and octs[i+1]!=-4):
        list_4.append(count_4)
        count_4=1