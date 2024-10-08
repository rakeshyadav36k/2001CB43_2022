# #Help https://youtu.be/H37f_x4wAC0
# def octant_longest_subsequence_count():
# ###Code

# from platform import python_version
# ver = python_version()

# if ver == "3.8.10":
#     print("Correct Version Installed")
# else:
#     print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")


# octant_longest_subsequence_count()


# I have imported some libraries

from openpyxl import load_workbook
import numpy as np
import pandas as pd

wb = load_workbook('input_octant_longest_subsequence_with_range.xlsx')     # I have opened the Excel file
sheet = wb.active

df = pd.read_excel('input_octant_longest_subsequence_with_range.xlsx')     # df is a data frame in which we put the data of excel file using pandas

sheet.cell(row=1, column=5).value = "U Avg"       # Written the header 
sheet.cell(row=1, column=6).value = "V Avg" 
sheet.cell(row=1, column=7).value = "W Avg"  
sheet.cell(row=1, column=8).value = "U'=U - U avg" 
sheet.cell(row=1, column=9).value = "V'=V - V avg" 
sheet.cell(row=1, column=10).value = "W'=W - W avg"

Uavg = df['U'].mean()     # calculated the mean using pandas
Vavg = df['V'].mean()
Wavg = df['W'].mean()

sheet.cell(row=2, column=5).value = Uavg   # added the average values in the sheet
sheet.cell(row=2, column=6).value = Vavg 
sheet.cell(row=2, column=7).value = Wavg 

l1 = df['U']            # creating the list l1,l2,l3 which consist the element of U,V,W respectively
l2 = df['V']
l3 = df['W']


# creating three lists l4,l5 & l6 which contains the values of U', V' & W'

#********************   
l4=[]
for i in l1:
    a = i - Uavg
    l4.append(a)

for i in range(2,len(l1)+2):
    sheet.cell(row=i, column=8).value = l4[i-2]

l5=[]
for i in l2:
    a = i - Vavg
    l5.append(a)

for i in range(2,len(l2)+2):
    sheet.cell(row=i, column=9).value = l5[i-2]

l6=[]
for i in l3:
    a = i - Wavg
    l6.append(a)

for i in range(2,len(l3)+2):
    sheet.cell(row=i, column=10).value = l6[i-2]

#************************************

# here we have created a header "Octant" and print the values of octants in excel file

sheet.cell(row=1, column=11).value = "Octant"



for i in range(0,len(l1)):
    if(l4[i]>0 and l5[i]>0):
        if(l6[i]>0):
            sheet.cell(row=i+2, column=11).value = "+1"
        else:                                               # this tells whether the octant is +1 or -1
            sheet.cell(row=i+2, column=11).value = "-1"
    elif(l4[i]<0 and l5[i]>0):
        if(l6[i]>0):
            sheet.cell(row=i+2, column=11).value = "+2"
        else:                                               # this tells whether the octant is +2 or -2
            sheet.cell(row=i+2, column=11).value = "-2"
    elif(l4[i]<0 and l5[i]<0):
        if(l6[i]>0):
            sheet.cell(row=i+2, column=11).value = "+3"
        else:                                                # this tells whether the octant is +3 or -3
            sheet.cell(row=i+2, column=11).value = "-3"
    elif(l4[i]>0 and l5[i]<0):
        if(l6[i]>0):
            sheet.cell(row=i+2, column=11).value = "+4"
        else:                                                 # this tells whether the octant is +4 or -4
            sheet.cell(row=i+2, column=11).value = "-4"
    

sheet.cell(row=1,column=13).value = "Octants"
sheet.cell(row=1,column=14).value = "Longest Sequence Length"
sheet.cell(row=1,column=15).value = "Count"

list_octant = ["+1","-1","+2","-2","+3","-3","+4","-4"]
for i in range(8):
    sheet.cell(row=2+i,column=13).value = list_octant[i]

octs=[]           # we put all the octants values in list octs
for i in range(len(l1)):
    x=sheet.cell(row=i+2,column=11).value
    octs.append(int(x))

time = []
for i in range(len(l1)):
    x=sheet.cell(row=i+2,column=1).value
    time.append(x)

row_no1 = []     #we have created this list to store the row number of the row form which a subsequence is starting. this row number is the index of list of time.
list1=[]      #list containing length of subsequents
count1=1
for i in range(len(octs)-1):
    if(octs[i]==1 and octs[i+1]==1):
        count1+=1
    elif(octs[i]==1 and octs[i+1]!=1):
        list1.append(count1)
        row_no1.append(i-count1+1)
        count1=1

time_l1 = []        #in this list we are collecting the range of time of a subsequence.
for i in range(len(list1)-1):
    if(list1[i]==max(list1)):
        a=row_no1[i]
        time_l1.append([time[a],time[a+max(list1)-1]])

row_no_1 = []          #we have created this list to store the row number of the row form which a subsequence is starting. this row number is the index of list of time.
list_1=[]      #list containing length of subsequents
count_1=1
for i in range(len(octs)-1):
    if(octs[i]==-1 and octs[i+1]==-1):
        count_1+=1
    elif(octs[i]==-1 and octs[i+1]!=-1):
        list_1.append(count_1)
        row_no_1.append(i-count_1+1)
        count_1=1

time_l_1 = []           #in this list we are collecting the range of time of a subsequence.
for i in range(len(list_1)-1):
    if(list_1[i]==max(list_1)):
        a=row_no_1[i]
        time_l_1.append([time[a],time[a+max(list_1)-1]])


row_no2 = []            #we have created this list to store the row number of the row form which a subsequence is starting. this row number is the index of list of time.
list2=[]      #list containing length of subsequents
count2=1
for i in range(len(octs)-1):
    if(octs[i]==2 and octs[i+1]==2):
        count2+=1
    elif(octs[i]==2 and octs[i+1]!=2):
        list2.append(count2)
        row_no2.append(i-count2+1)
        count2=1

time_l2 = []          #in this list we are collecting the range of time of a subsequence.
for i in range(len(list2)-1):
    if(list2[i]==max(list2)):
        a=row_no2[i]
        time_l2.append([time[a],time[a+max(list2)-1]])

row_no_2 = []       #we have created this list to store the row number of the row form which a subsequence is starting. this row number is the index of list of time.
list_2=[]      #list containing length of subsequents
count_2=1
for i in range(len(octs)-1):
    if(octs[i]==-2 and octs[i+1]==-2):
        count_2+=1
    elif(octs[i]==-2 and octs[i+1]!=-2):
        list_2.append(count_2)
        row_no_2.append(i-count_2+1)
        count_2=1

time_l_2 = []         #in this list we are collecting the range of time of a subsequence.
for i in range(len(list_2)-1):
    if(list_2[i]==max(list_2)):
        a=row_no_2[i]
        time_l_2.append([time[a],time[a+max(list_2)-1]])

row_no3 = []           #we have created this list to store the row number of the row form which a subsequence is starting. this row number is the index of list of time.
list3=[]      #list containing length of subsequents
count3=1
for i in range(len(octs)-1):
    if(octs[i]==3 and octs[i+1]==3):
        count3+=1
    elif(octs[i]==3 and octs[i+1]!=3):
        list3.append(count3)
        row_no3.append(i-count3+1)
        count3=1

time_l3 = []           #in this list we are collecting the range of time of a subsequence.
for i in range(len(list3)-1):
    if(list3[i]==max(list3)):
        a=row_no3[i]
        time_l3.append([time[a],time[a+max(list3)-1]])

row_no_3 = []          #we have created this list to store the row number of the row form which a subsequence is starting. this row number is the index of list of time.
list_3=[]      #list containing length of subsequents
count_3=1
for i in range(len(octs)-1):
    if(octs[i]==-3 and octs[i+1]==-3):
        count_3+=1
    elif(octs[i]==-3 and octs[i+1]!=-3):
        list_3.append(count_3)
        row_no_3.append(i-count_3+1)
        count_3=1

time_l_3 = []             #in this list we are collecting the range of time of a subsequence.
for i in range(len(list_3)-1):
    if(list_3[i]==max(list_3)):
        a=row_no_3[i]
        time_l_3.append([time[a],time[a+max(list_3)-1]])

row_no4 = []          #we have created this list to store the row number of the row form which a subsequence is starting. this row number is the index of list of time.
list4=[]      #list containing length of subsequents
count4=1
for i in range(len(octs)-1):
    if(octs[i]==4 and octs[i+1]==4):
        count4+=1
    elif(octs[i]==4 and octs[i+1]!=4):
        list4.append(count4)
        row_no4.append(i-count4+1)
        count4=1
 
time_l4 = []            #in this list we are collecting the range of time of a subsequence.
for i in range(len(list4)-1):
    if(list4[i]==max(list4)):
        a=row_no4[i]
        time_l4.append([time[a],time[a+max(list4)-1]])

row_no_4 = []        #we have created this list to store the row number of the row form which a subsequence is starting. this row number is the index of list of time.
list_4=[]      #list containing length of subsequents
count_4=1
for i in range(len(octs)-1):
    if(octs[i]==-4 and octs[i+1]==-4):
        count_4+=1
    elif(octs[i]==-4 and octs[i+1]!=-4):
        list_4.append(count_4)
        row_no_4.append(i-count_4+1)
        count_4=1

time_l_4 = []              #in this list we are collecting the range of time of a subsequence.
for i in range(len(list_4)-1):
    if(list_4[i]==max(list_4)):
        a=row_no_4[i]
        time_l_4.append([time[a],time[a+max(list_4)-1]])

# here we have printed the longest subsequence of octants
sheet['N2'] = max(list1)
sheet['N3'] = max(list_1)
sheet['N4'] = max(list2)
sheet['N5'] = max(list_2)
sheet['N6'] = max(list3)
sheet['N7'] = max(list_3)
sheet['N8'] = max(list4)
sheet['N9'] = max(list_4)

# here we have printed the howmany times maximum subsequence have occur
sheet['O2'] = list1.count(max(list1))
sheet['O3'] = list_1.count(max(list_1))
sheet['O4'] = list2.count(max(list2))
sheet['O5'] = list_2.count(max(list_2))
sheet['O6'] = list3.count(max(list3))
sheet['O7'] = list_3.count(max(list_3))
sheet['O8'] = list4.count(max(list4))
sheet['O9'] = list_4.count(max(list_4))

sheet.cell(row=1,column=17).value = "Octant"
sheet.cell(row=1,column=18).value = "Longest Subsequence Length"
sheet.cell(row=1,column=19).value = "Count"

sheet.cell(row=3,column=17).value = "Time"     # here we have written the table
sheet.cell(row=3,column=18).value = "From"
sheet.cell(row=3,column=19).value = "To"
sheet.cell(row=6,column=17).value = "Time"
sheet.cell(row=6,column=18).value = "From"
sheet.cell(row=6,column=19).value = "To"
sheet.cell(row=11,column=17).value = "Time"
sheet.cell(row=11,column=18).value = "From"
sheet.cell(row=11,column=19).value = "To"
sheet.cell(row=14,column=17).value = "Time"
sheet.cell(row=14,column=18).value = "From"
sheet.cell(row=14,column=19).value = "To"
sheet.cell(row=17,column=17).value = "Time"
sheet.cell(row=17,column=18).value = "From"
sheet.cell(row=17,column=19).value = "To"
sheet.cell(row=20,column=17).value = "Time"
sheet.cell(row=20,column=18).value = "From"
sheet.cell(row=20,column=19).value = "To"
sheet.cell(row=23,column=17).value = "Time"
sheet.cell(row=23,column=18).value = "From"
sheet.cell(row=23,column=19).value = "To"
sheet.cell(row=26,column=17).value = "Time"
sheet.cell(row=26,column=18).value = "From"
sheet.cell(row=26,column=19).value = "To"


sheet.cell(row=2,column=17).value = "+1"          # we are putting the values of length of sequence, maximum count & time range of subsequence of "+1"
sheet.cell(row=2,column=18).value = max(list1)
sheet.cell(row=2,column=19).value = list1.count(max(list1))
sheet.cell(row=4,column=18).value = time_l1[0][0]
sheet.cell(row=4,column=19).value = time_l1[0][1]

sheet.cell(row=5,column=17).value = "-1"            # we are putting the values of length of sequence, maximum count & time range of subsequence of "-1"
sheet.cell(row=5,column=18).value = max(list_1)
sheet.cell(row=5,column=19).value = list_1.count(max(list_1))
sheet.cell(row=7,column=18).value = time_l_1[0][0]
sheet.cell(row=7,column=19).value = time_l_1[0][1]
sheet.cell(row=8,column=18).value = time_l_1[1][0]
sheet.cell(row=8,column=19).value = time_l_1[1][1]
sheet.cell(row=9,column=18).value = time_l_1[2][0]
sheet.cell(row=9,column=19).value = time_l_1[2][1]

sheet.cell(row=10,column=17).value = "+2"       # we are putting the values of length of sequence, maximum count & time range of subsequence of "+2"
sheet.cell(row=10,column=18).value = max(list2)
sheet.cell(row=10,column=19).value = list2.count(max(list2))
sheet.cell(row=12,column=18).value = time_l2[0][0]
sheet.cell(row=12,column=19).value = time_l2[0][1]

sheet.cell(row=13,column=17).value = "-2"       # we are putting the values of length of sequence, maximum count & time range of subsequence of "-2"
sheet.cell(row=13,column=18).value = max(list_2)
sheet.cell(row=13,column=19).value = list_2.count(max(list_2))
sheet.cell(row=15,column=18).value = time_l_2[0][0]
sheet.cell(row=15,column=19).value = time_l_2[0][1]

sheet.cell(row=16,column=17).value = "+3"         # we are putting the values of length of sequence, maximum count & time range of subsequence of "+3"
sheet.cell(row=16,column=18).value = max(list3)
sheet.cell(row=16,column=19).value = list3.count(max(list3))
sheet.cell(row=18,column=18).value = time_l3[0][0]
sheet.cell(row=18,column=19).value = time_l3[0][1]

sheet.cell(row=19,column=17).value = "-3"     # we are putting the values of length of sequence, maximum count & time range of subsequence of "-3"
sheet.cell(row=19,column=18).value = max(list_3)
sheet.cell(row=19,column=19).value = list_3.count(max(list_3))
sheet.cell(row=21,column=18).value = time_l_3[0][0]
sheet.cell(row=21,column=19).value = time_l_3[0][1]

sheet.cell(row=22,column=17).value = "+4"       # we are putting the values of length of sequence, maximum count & time range of subsequence of "+4"
sheet.cell(row=22,column=18).value = max(list4)
sheet.cell(row=22,column=19).value = list4.count(max(list4))
sheet.cell(row=24,column=18).value = time_l4[0][0]
sheet.cell(row=24,column=19).value = time_l4[0][1]

sheet.cell(row=25,column=17).value = "-4"         # we are putting the values of length of sequence, maximum count & time range of subsequence of "-4"
sheet.cell(row=25,column=18).value = max(list_4)
sheet.cell(row=25,column=19).value = list_4.count(max(list_4))
sheet.cell(row=27,column=18).value = time_l_4[0][0]
sheet.cell(row=27,column=19).value = time_l_4[0][1]





wb.save('output_octant_longest_subsequence_with_range_2001CB43.xlsx')