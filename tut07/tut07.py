# imported some libraries
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill
import numpy as np
import pandas as pd
import os
from datetime import datetime
from openpyxl.utils.dataframe import dataframe_to_rows

start=datetime.now()

filenames = []
path = 'input'
path_out = 'output'
cwd = os.getcwd()                   # this code is used for taking multiple input file
os.chdir('input')
for file in os.listdir():
    file_name = os.path.basename(file)
    file_name = os.path.splitext(file_name)[0]
    filenames.append(file_name)

    wb = load_workbook(file)     # I have opened the Excel file
    sheet = wb.worksheets[0]

    df = pd.read_excel(file)     # df is a data frame in which we put the data of excel file using pandas


    top = Side(border_style='thin',color="000000")            # this whole code is for the border of a shell
    bottom = Side(border_style='thin',color="000000")
    left = Side(border_style='thin',color="000000")
    right = Side(border_style='thin',color="000000")

    border = Border(top=top,bottom=bottom,left=left,right=right)


    pattern = PatternFill(patternType='solid', fgColor="FFFF00")   # this code is for yellow color
    

    sheet.cell(row=1, column=5).value = "U Avg"       # Written the header 
    sheet.cell(row=1, column=6).value = "V Avg" 
    sheet.cell(row=1, column=7).value = "W Avg"  
    sheet.cell(row=1, column=8).value = "U'=U - U avg" 
    sheet.cell(row=1, column=9).value = "V'=V - V avg" 
    sheet.cell(row=1, column=10).value = "W'=W - W avg"

    Uavg = df['U'].mean()     # calculated the mean using pandas
    Vavg = df['V'].mean()
    Wavg = df['W'].mean()

    sheet.cell(row=2, column=5).value = Uavg   # added the average values in the sheet
    sheet.cell(row=2, column=6).value = Vavg 
    sheet.cell(row=2, column=7).value = Wavg 

    l1 = df['U']            # creating the list l1,l2,l3 which consist the element of U,V,W respectively
    l2 = df['V']
    l3 = df['W']


    # creating three lists l4,l5 & l6 which contains the values of U', V' & W'

    #********************   
    l4=[]
    for i in l1:
        a = i - Uavg
        l4.append(a)

    for i in range(2,len(l1)+2):
        sheet.cell(row=i, column=8).value = l4[i-2]

    l5=[]
    for i in l2:
        a = i - Vavg
        l5.append(a)

    for i in range(2,len(l2)+2):
        sheet.cell(row=i, column=9).value = l5[i-2]

    l6=[]
    for i in l3:
        a = i - Wavg
        l6.append(a)

    for i in range(2,len(l3)+2):
        sheet.cell(row=i, column=10).value = l6[i-2]

    #************************************

    # here we have created a header "Octant" and print the values of octants in excel file

    sheet.cell(row=1, column=11).value = "Octant"



    for i in range(0,len(l1)):
        if(l4[i]>0 and l5[i]>0):
            if(l6[i]>0):
                sheet.cell(row=i+2, column=11).value = "+1"
            else:                                               # this tells whether the octant is +1 or -1
                sheet.cell(row=i+2, column=11).value = "-1"
        elif(l4[i]<0 and l5[i]>0):
            if(l6[i]>0):
                sheet.cell(row=i+2, column=11).value = "+2"
            else:                                               # this tells whether the octant is +2 or -2
                sheet.cell(row=i+2, column=11).value = "-2"
        elif(l4[i]<0 and l5[i]<0):
            if(l6[i]>0):
                sheet.cell(row=i+2, column=11).value = "+3"
            else:                                                # this tells whether the octant is +3 or -3
                sheet.cell(row=i+2, column=11).value = "-3"
        elif(l4[i]>0 and l5[i]<0):
            if(l6[i]>0):
                sheet.cell(row=i+2, column=11).value = "+4"
            else:                                                 # this tells whether the octant is +4 or -4
                sheet.cell(row=i+2, column=11).value = "-4"


    # this list l7 contains the all octants values
    l7=[]
    for i in range(len(l1)):
        x=sheet.cell(row=i+2,column=11).value
        l7.append(int(x))


    sheet['N1']="Overall Octant Count"  # this is basicallly printed the header
    sheet['N3']="Octant ID"  
    sheet['N3'].border = border  
    sheet['N4']="Overall count"  
    sheet['N4'].border = border  
    sheet['O3'] = "+1"
    sheet['O3'].border = border
    sheet['P3'] = "-1"
    sheet['P3'].border = border
    sheet['Q3'] = "+2"
    sheet['Q3'].border = border
    sheet['R3'] = "-2"
    sheet['R3'].border = border
    sheet['S3'] = "+3"
    sheet['S3'].border = border
    sheet['T3'] = "-3"
    sheet['T3'].border = border
    sheet['U3'] = "+4"
    sheet['U3'].border = border
    sheet['V3'] = "-4"
    sheet['V3'].border = border

    ctpos1 = ctneg1 = ctpos2 = ctneg2 = ctpos3 = ctneg3 = ctpos4 = ctneg4 = 0  # these variables are total no each octant present

    for i in range(0,len(l1)):
        if(l4[i]>0 and l5[i]>0):
            if(l6[i]>0):
                ctpos1 += 1               # total count of octant no +1 & -1
            else:
                ctneg1 += 1
        elif(l4[i]<0 and l5[i]>0):
            if(l6[i]>0):
                ctpos2 += 1                 # total count of octant no +2 & -2
            else:
                ctneg2 += 1
        elif(l4[i]<0 and l5[i]<0):
            if(l6[i]>0):
                ctpos3 += 1                 # total count of octant no +3 & -3
            else:
                ctneg3 += 1
        elif(l4[i]>0 and l5[i]<0):
            if(l6[i]>0):
                ctpos4 += 1                 # total count of octant no +4 & -4
            else:
                ctneg4 += 1


    #  we have inserted the values of total no of each count

    sheet.cell(row=4, column=15).value = ctpos1
    sheet.cell(row=4, column=15).border = border
    sheet.cell(row=4, column=16).value = ctneg1
    sheet.cell(row=4, column=16).border = border
    sheet.cell(row=4, column=17).value = ctpos2
    sheet.cell(row=4, column=17).border = border
    sheet.cell(row=4, column=18).value = ctneg2
    sheet.cell(row=4, column=18).border = border
    sheet.cell(row=4, column=19).value = ctpos3
    sheet.cell(row=4, column=19).border = border
    sheet.cell(row=4, column=20).value = ctneg3
    sheet.cell(row=4, column=20).border = border
    sheet.cell(row=4, column=21).value = ctpos4
    sheet.cell(row=4, column=21).border = border
    sheet.cell(row=4, column=22).value = ctneg4
    sheet.cell(row=4, column=22).border = border


    mod=5000    # this is a user defined mod value
    if(len(l7)%mod!=0):           
        p = len(l7)//mod + 1           #variable p is no of partitions
    else:
        p = len(l7)//mod 

    sheet['M4'] = "Mod" + " " +str(mod)

    A = []              # here we have taken a list A which contains another list B(list B contains the octants values of partition)
    x=0
    new_mod = mod
    for i in range(p):
        B = []
        for j in range(x,x + new_mod):
            B.append(l7[j])
        x+=mod
        if((len(l7)-x)<mod):
                new_mod = len(l7) - x
        A.append(B)


    new_octant = [1,-1,2,-2,3,-3,4,-4]   #this is a list of octant values
    for i in range(p):
        if(mod*(i+1)<len(l7)):
            sheet.cell(row=i+5, column=14).value = str(mod*i)+"-"+str(mod*(i+1)-1)
            sheet.cell(row=i+5, column=14).border = border
        else:
            sheet.cell(row=i+5, column=14).value = str(mod*i)+"-"+str(len(l7)-1)
            sheet.cell(row=i+5, column=14).border = border
        for j in range(8):
            sheet.cell(row=i+5, column=15+j).value = A[i].count(new_octant[j])
            sheet.cell(row=i+5, column=15+j).border = border          #we have counted the octant values in a range mod value

    ################################
    sheet.cell(row= 1,  column=35).value = "Overall Transition Count"    # it is simply written the headers
    sheet.cell(row=3, column=35).value = "Octant #"
    sheet.cell(row=3, column=35).border = border

    octant = ["+1","-1","+2","-2","+3","-3","+4","-4"]
    for i in range(8):
        sheet.cell(row=4+i, column=35).value = octant[i]
        sheet.cell(row=3, column=36+i).value = octant[i]
        sheet.cell(row=4+i, column=35).border = border
        sheet.cell(row=3, column=36+i).border = border

    sheet.cell(row=4, column=34).value = "From"
    sheet.cell(row=2, column=36).value = "To"

    # now are defining the transition from one octants to another octant using 64 variables 

    r1c1 = r1c2 = r1c3 = r1c4 = r1c5 = r1c6 = r1c7 = r1c8 = 0
    r2c1 = r2c2 = r2c3 = r2c4 = r2c5 = r2c6 = r2c7 = r2c8 = 0
    r3c1 = r3c2 = r3c3 = r3c4 = r3c5 = r3c6 = r3c7 = r3c8 = 0
    r4c1 = r4c2 = r4c3 = r4c4 = r4c5 = r4c6 = r4c7 = r4c8 = 0
    r5c1 = r5c2 = r5c3 = r5c4 = r5c5 = r5c6 = r5c7 = r5c8 = 0
    r6c1 = r6c2 = r6c3 = r6c4 = r6c5 = r6c6 = r6c7 = r6c8 = 0
    r7c1 = r7c2 = r7c3 = r7c4 = r7c5 = r7c6 = r7c7 = r7c8 = 0
    r8c1 = r8c2 = r8c3 = r8c4 = r8c5 = r8c6 = r8c7 = r8c8 = 0 

    for i in range(len(l7)-1):      # these codes are to count the each transition values
        if(l7[i]==1):
            if(l7[i+1]==1):
                r1c1+=1
            elif(l7[i+1]==-1):
                r1c2+=1
            elif(l7[i+1]==2):
                r1c3+=1
            elif(l7[i+1]==-2):
                r1c4+=1
            elif(l7[i+1]==3):
                r1c5+=1
            elif(l7[i+1]==-3):
                r1c6+=1
            elif(l7[i+1]==4):
                r1c7+=1
            elif(l7[i+1]==-4):
                r1c8+=1

        if(l7[i]==-1):
            if(l7[i+1]==1):
                r2c1+=1
            elif(l7[i+1]==-1):
                r2c2+=1
            elif(l7[i+1]==2):
                r2c3+=1
            elif(l7[i+1]==-2):
                r2c4+=1
            elif(l7[i+1]==3):
                r2c5+=1
            elif(l7[i+1]==-3):
                r2c6+=1
            elif(l7[i+1]==4):
                r2c7+=1
            elif(l7[i+1]==-4):
                r2c8+=1

        if(l7[i]==2):
            if(l7[i+1]==1):
                r3c1+=1
            elif(l7[i+1]==-1):
                r3c2+=1
            elif(l7[i+1]==2):
                r3c3+=1
            elif(l7[i+1]==-2):
                r3c4+=1
            elif(l7[i+1]==3):
                r3c5+=1
            elif(l7[i+1]==-3):
                r3c6+=1
            elif(l7[i+1]==4):
                r3c7+=1
            elif(l7[i+1]==-4):
                r3c8+=1

        if(l7[i]==-2):
            if(l7[i+1]==1):
                r4c1+=1
            elif(l7[i+1]==-1):
                r4c2+=1
            elif(l7[i+1]==2):
                r4c3+=1
            elif(l7[i+1]==-2):
                r4c4+=1
            elif(l7[i+1]==3):
                r4c5+=1
            elif(l7[i+1]==-3):
                r4c6+=1
            elif(l7[i+1]==4):
                r4c7+=1
            elif(l7[i+1]==-4):
                r4c8+=1

        if(l7[i]==3):
            if(l7[i+1]==1):
                r5c1+=1
            elif(l7[i+1]==-1):
                r5c2+=1
            elif(l7[i+1]==2):
                r5c3+=1
            elif(l7[i+1]==-2):
                r5c4+=1
            elif(l7[i+1]==3):
                r5c5+=1
            elif(l7[i+1]==-3):
                r5c6+=1
            elif(l7[i+1]==4):
                r5c7+=1
            elif(l7[i+1]==-4):
                r5c8+=1

        if(l7[i]==-3):
            if(l7[i+1]==1):
                r6c1+=1
            elif(l7[i+1]==-1):
                r6c2+=1
            elif(l7[i+1]==2):
                r6c3+=1
            elif(l7[i+1]==-2):
                r6c4+=1
            elif(l7[i+1]==3):
                r6c5+=1
            elif(l7[i+1]==-3):
                r6c6+=1
            elif(l7[i+1]==4):
                r6c7+=1
            elif(l7[i+1]==-4):
                r6c8+=1

        if(l7[i]==4):
            if(l7[i+1]==1):
                r7c1+=1
            elif(l7[i+1]==-1):
                r7c2+=1
            elif(l7[i+1]==2):
                r7c3+=1
            elif(l7[i+1]==-2):
                r7c4+=1
            elif(l7[i+1]==3):
                r7c5+=1
            elif(l7[i+1]==-3):
                r7c6+=1
            elif(l7[i+1]==4):
                r7c7+=1
            elif(l7[i+1]==-4):
                r7c8+=1

        if(l7[i]==-4):
            if(l7[i+1]==1):
                r8c1+=1
            elif(l7[i+1]==-1):
                r8c2+=1
            elif(l7[i+1]==2):
                r8c3+=1
            elif(l7[i+1]==-2):
                r8c4+=1
            elif(l7[i+1]==3):
                r8c5+=1
            elif(l7[i+1]==-3):
                r8c6+=1
            elif(l7[i+1]==4):
                r8c7+=1
            elif(l7[i+1]==-4):
                r8c8+=1

        # here transition values are put in the list by which it will easily print
    l8 = [r1c1,r1c2,r1c3,r1c4,r1c5,r1c6,r1c7,r1c8]
    l9 = [r2c1,r2c2,r2c3,r2c4,r2c5,r2c6,r2c7,r2c8]
    l10 = [r3c1,r3c2,r3c3,r3c4,r3c5,r3c6,r3c7,r3c8]
    l11 = [r4c1,r4c2,r4c3,r4c4,r4c5,r4c6,r4c7,r4c8]
    l12 = [r5c1,r5c2,r5c3,r5c4,r5c5,r5c6,r5c7,r5c8]
    l13 = [r6c1,r6c2,r6c3,r6c4,r6c5,r6c6,r6c7,r6c8]
    l14 = [r7c1,r7c2,r7c3,r7c4,r7c5,r7c6,r7c7,r7c8]
    l15 = [r8c1,r8c2,r8c3,r8c4,r8c5,r8c6,r8c7,r8c8]

    for i in range(8):                    # here have printed the overall transition values
        sheet.cell(row=4,column=36+i).value = l8[i]
        sheet.cell(row=5,column=36+i).value = l9[i]
        sheet.cell(row=6,column=36+i).value = l10[i]
        sheet.cell(row=7,column=36+i).value = l11[i]
        sheet.cell(row=8,column=36+i).value = l12[i]
        sheet.cell(row=9,column=36+i).value = l13[i]
        sheet.cell(row=10,column=36+i).value = l14[i]
        sheet.cell(row=11,column=36+i).value = l15[i]

    for i in range(8):                    # here have bordered the overall transition values
        sheet.cell(row=4,column=36+i).border = border
        sheet.cell(row=5,column=36+i).border = border
        sheet.cell(row=6,column=36+i).border = border
        sheet.cell(row=7,column=36+i).border = border
        sheet.cell(row=8,column=36+i).border = border
        sheet.cell(row=9,column=36+i).border = border
        sheet.cell(row=10,column=36+i).border = border
        sheet.cell(row=11,column=36+i).border = border
